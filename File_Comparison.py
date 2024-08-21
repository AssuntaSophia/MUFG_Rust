# -*- coding: utf-8 -*-
"""
Created on Wed Aug 21 10:43:13 2024

@author: User
"""

#%%
import tkinter as tk
from tkinter import messagebox, filedialog
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from tkinter import Tk, Button, Label, messagebox, StringVar, Entry
from tkinter.filedialog import askopenfilename
import os
import subprocess

# Function to highlight differences in the second workbook
def highlight_differences(file1, file2, sheet_name=None, compare_type="normal"):
    wb2 = load_workbook(file2)
    
    if compare_type == "high_low":
        fill_higher = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green
        fill_lower = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")   # Red
    else:
        fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow

    if sheet_name:
        sheets = [sheet_name]
    else:
        sheets = wb2.sheetnames

    with pd.ExcelFile(file1) as xls1:
        for sheet in sheets:
            if sheet in xls1.sheet_names:
                df1 = pd.read_excel(file1, sheet_name=sheet)
                df2 = pd.read_excel(file2, sheet_name=sheet)
                ws = wb2[sheet]

                # Ensure number of rows and columns are within valid range
                max_rows = ws.max_row
                max_cols = ws.max_column

                for row in range(1, max_rows + 1):
                    for col in range(1, max_cols + 1):
                        if row <= df1.shape[0] and col <= df1.shape[1]:
                            cell1 = df1.iat[row-1, col-1]
                            cell2 = df2.iat[row-1, col-1]
                            if pd.isna(cell1):
                                continue  # Skip blank cells in file1
                            if compare_type == "high_low":
                                if cell1 < cell2:
                                    excel_cell = ws.cell(row=row, column=col)
                                    excel_cell.fill = fill_higher
                                elif cell1 > cell2:
                                    excel_cell = ws.cell(row=row, column=col)
                                    excel_cell.fill = fill_lower
                            else:
                                if cell1 != cell2:
                                    excel_cell = ws.cell(row=row, column=col)
                                    excel_cell.fill = fill

    comparison_file = "COMPARISON.xlsx"
    wb2.save(comparison_file)
    subprocess.Popen([comparison_file], shell=True)

# Function to select files and perform comparison
def compare_all_sheets():
    global file1, file2, compare_all_var, sheet_name_entry

    if not file1 or not file2:
        messagebox.showerror("Error", "Please select both Excel files.")
        return

    compare_all = compare_all_var.get()
    if compare_all:
        highlight_differences(file1, file2)
    else:
        sheet_name = sheet_name_entry.get()
        if not sheet_name:
            messagebox.showerror("Error", "Please enter a sheet name.")
            return
        highlight_differences(file1, file2, sheet_name)

# Function to open window for entering sheet name
def open_sheet_name_window():
    global sheet_name_entry, sheet_name_window
    sheet_name_window = tk.Toplevel(root)
    sheet_name_window.title("Enter Sheet Name")
    sheet_name_window.geometry("300x150")
    sheet_name_window.configure(bg="#007BFF")  # tandard application blue
    
    # Label and Entry for sheet name
    tk.Label(sheet_name_window, text="Enter Sheet Name:", font=("Roboto", 12), bg="#007BFF", fg="white").pack(pady=10)
    sheet_name_entry = tk.Entry(sheet_name_window, width=30, font=("Roboto", 10))
    sheet_name_entry.pack(pady=5)
    
    # Button to start comparison
    tk.Button(sheet_name_window, text="Compare", command=perform_specific_comparison, bg="#3498DB", fg="white", font=("Roboto", 10)).pack(pady=10)

# Function to perform comparison for specific sheet
def perform_specific_comparison():
    global sheet_name_window
    sheet_name = sheet_name_entry.get().strip()

    if not sheet_name:
        messagebox.showerror("Error", "Please enter a sheet name.")
        return

    highlight_differences(file1, file2, sheet_name)

    # Close the sheet name window after comparison
    sheet_name_window.destroy()

# Function to handle credits button click
def show_credits():
    messagebox.showinfo("Credits", "This application was created by Assunta Felice.")

# Function to select the first Excel file
def select_file1():
    global file1
    file1 = filedialog.askopenfilename(title="Select the first Excel file", filetypes=[("Excel files", "*.xlsx")])
    if file1:
        label_file1.config(text=f"Selected: {file1}")

# Function to select the second Excel file
def select_file2():
    global file2
    file2 = filedialog.askopenfilename(title="Select the second Excel file", filetypes=[("Excel files", "*.xlsx")])
    if file2:
        label_file2.config(text=f"Selected: {file2}")

# Function to perform high/low comparison
def high_low_comparison():
    global file1, file2, compare_all_var, sheet_name_entry

    if not file1 or not file2:
        messagebox.showerror("Error", "Please select both Excel files.")
        return

    compare_all = compare_all_var.get()
    if compare_all:
        highlight_differences(file1, file2, compare_type="high_low")
    else:
        sheet_name = sheet_name_entry.get()
        if not sheet_name:
            messagebox.showerror("Error", "Please enter a sheet name.")
            return
        highlight_differences(file1, file2, sheet_name, compare_type="high_low")

# Initialize the main Tkinter window
root = tk.Tk()
root.title("Excel File Comparison")
root.geometry("400x300")
root.configure(bg="#007BFF")  # More standard application blue

file1 = None
file2 = None
compare_all_var = StringVar(value="all")

# Title label
title_label = tk.Label(root, text="Excel File Comparison", font=("Roboto", 18), bg="#007BFF", fg="white")
title_label.pack(pady=10)

# Credits button
credits_button = tk.Button(root, text="Credits", command=show_credits, bg="#3498DB", fg="white", font=("Roboto", 10))
credits_button.pack()

# Frame for file selection buttons
file_frame = tk.Frame(root, bg="#007BFF", bd=2, relief=tk.RAISED)
file_frame.pack(pady=20)

# File selection buttons
tk.Label(file_frame, text="File 1:", font=("Roboto", 12), bg="#007BFF", fg="white").grid(row=0, column=0, padx=10, pady=5)
btn_file1 = tk.Button(file_frame, text="Select Excel File", command=select_file1, bg="#3498DB", fg="white", font=("Roboto", 10))
btn_file1.grid(row=0, column=1, padx=10, pady=5)

label_file1 = tk.Label(file_frame, text="No file selected", bg="#007BFF", fg="white", font=("Roboto", 10))
label_file1.grid(row=0, column=2, padx=10, pady=5)

tk.Label(file_frame, text="File 2:", font=("Roboto", 12), bg="#007BFF", fg="white").grid(row=1, column=0, padx=10, pady=5)
btn_file2 = tk.Button(file_frame, text="Select Excel File", command=select_file2, bg="#3498DB", fg="white", font=("Roboto", 10))
btn_file2.grid(row=1, column=1, padx=10, pady=5)

label_file2 = tk.Label(file_frame, text="No file selected", bg="#007BFF", fg="white", font=("Roboto", 10))
label_file2.grid(row=1, column=2, padx=10, pady=5)

# Frame for checkbox and compare individual sheet button
compare_frame = tk.Frame(root, bg="#007BFF")
compare_frame.pack(pady=10)

# Checkbox for comparing all sheets
compare_all_var = tk.BooleanVar()
compare_all_checkbox = tk.Checkbutton(compare_frame, text="Compare All Sheets", variable=compare_all_var, bg="#007BFF", font=("Roboto", 10))
compare_all_checkbox.pack(side=tk.LEFT)

# Button to open window for entering sheet name (next to the checkbox)
btn_compare_individual = tk.Button(compare_frame, text="Compare Individual Sheet", command=open_sheet_name_window, bg="#3498DB", fg="white", font=("Roboto", 10))
btn_compare_individual.pack(side=tk.LEFT, padx=10)

# Button to compare all sheets (underneath the above frame)
btn_compare_all = tk.Button(root, text="Compare All Sheets", command=compare_all_sheets, bg="#3498DB", fg="white", font=("Roboto", 10))
btn_compare_all.pack(pady=5)

# Button to perform high/low comparison
btn_high_low = tk.Button(root, text="High/Low", command=high_low_comparison, bg="#3498DB", fg="white", font=("Roboto", 10))
btn_high_low.pack(pady=5)

# Run the Tkinter event loop
root.mainloop()
