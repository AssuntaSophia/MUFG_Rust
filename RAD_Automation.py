# -*- coding: utf-8 -*-
"""
Created on Wed Aug 21 11:02:13 2024

@author: User
"""

#imports
import calendar
import datetime
import os
import sys
import shutil
import openpyxl
import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
import math

#%%

def last_business_day():
    """Finds the last business day of the previous month in the format 
    01Jan24"""
    # Get the current date
    today = datetime.date.today()
    
    # Get the year and month of the previous month
    if today.month == 1:
        prev_month_year = today.year - 1
        prev_month = 12
    else:
        prev_month_year = today.year
        prev_month = today.month - 1
    
    # Get the last day of the previous month
    last_day_prev_month = calendar.monthrange(prev_month_year, prev_month)[1]
    
    # Iterate from the last day backwards until we find a business day
    for day in range(last_day_prev_month, 0, -1):
        date = datetime.date(prev_month_year, prev_month, day)
        if date.weekday() < 5:  # Monday to Friday (0 to 4)
            return date.strftime('%d%b%y')
        
    return None  # If no business day found in the previous month


    
def copy_file(source, destination):
    """Copies a file from a source to a destination"""
    os.system(f'copy "{source}" "{destination}"')
    
    
def Create_Dated_RAD(base_directory):
    """Function creates the Dated RAD in the Dated Archive Folder"""
    
    # Get the last business day of the last month
    business_day = last_business_day()
    if not last_business_day:
        print("Error: Unable to determine last business day of previous month.")
        return
    
    # Check if the files in the "Dated Archive" folder contain the last business day as a suffix
    dated_archive_folder = os.path.join(base_directory, "Dated Archive")
    
    files_in_dated_archive = os.listdir(dated_archive_folder)
    if any(file_name.endswith('_' + business_day) for file_name in files_in_dated_archive):
        print("Files with last business day suffix already exist in Dated Archive folder.")
        return
           
     # Copy files with suffixes "Muse" and "Bank" from the base directory to the Dated Archive folder
    muse_file = os.path.join(base_directory, "RARCPMonitoring_MUSE.xlsx")
    bank_file = os.path.join(base_directory, "RARCPMonitoring_Bank.xlsx")
    


    if os.path.exists(muse_file) and os.path.exists(bank_file):
        new_muse_file = os.path.join(dated_archive_folder, f"RARCPMonitoring_MUSE_{business_day}.xlsx")
        new_bank_file = os.path.join(dated_archive_folder, f"RARCPMonitoring_Bank_{business_day}.xlsx")
        
        copy_file(muse_file, new_muse_file)
        copy_file(bank_file, new_bank_file)

        print("Files copied to Dated Archive folder with last business day suffix.")
        return (new_muse_file, new_bank_file)
    else:
        print("Error: Files with suffixes Muse and Bank not found in the base directory.")
        return None      
    
    
 

def Locate_Requests(base_directory):
    """Locates the request in the Supporting Info Sub Folder"""
    # Get the current month
    today = datetime.date.today()
    month_name = today.strftime('%B')
    
    month_number = today.month
    
    # Format the current month with 'number.name' format
    month_suffix = f"{month_number}.{month_name}"
    
    #go into the supporting info sub-folder
    supp_info = os.path.join(base_directory, "Supporting info")
    
    # Add the relevant year folder onto the directory
    year_folder = os.path.join(supp_info, str(today.year))
    
    # Navigate to the RADs folder
    rads_folder = os.path.join(year_folder, 'RADs')
    
    # Open the folder based on the current month
    month_folder = os.path.join(rads_folder, month_suffix)
    print(month_folder)
    # Check if the folder exists
    if os.path.exists(month_folder):
        return month_folder
    else:
        print("Folder for the current month not found.")
        return None    
    
    
     
    
def process_request_file(request_file, request_directory):
    """Create a folder for all the completed requests
    Move requests to thsi folder"""
    
    folder_name = "Completed_Requests"
    folder_path = os.path.join(request_directory, folder_name)
    
    # Create the Completed_Requests folder if it doesn't exist
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)
    
    # Move the processed file to the "Completed_Requests" directory
    shutil.move(request_file, os.path.join(folder_path, os.path.basename(request_file)))


def process_all_files(request_directory):
    """Move all the files to the Completed Request Folder"""
    
    # Locate all files in the request directory
    all_files = [f for f in os.listdir(request_directory) if os.path.isfile(os.path.join(request_directory, f))]
    
    for file in all_files:
        file_path = os.path.join(request_directory, file)
        process_request_file(file_path, request_directory)    
    

def Code_Translate(Tech_Code):
    """Maps the Code for the Technology Inputs to the RAD KRIs"""
    
    code_mapping = {
            150: 1865, 175:1833, 177:1834,
            179:1847, 220:1606, 221:1605,
            183:1837, 185:1795, 187:1838,
            188:1839, 193:1848 
            }
    
    if Tech_Code in code_mapping:
        return code_mapping[Tech_Code]
    else:
        return None


#%%
def scale_to_order_of_magnitude(input_value, reference_value):
    """
    Scales the input value to match the order of magnitude of the reference value.

    Parameters:
    - input_value: The value to be scaled.
    - reference_value: The reference value to determine the desired order of magnitude.

    Returns:
    - The scaled value.
    """
    
    if input_value is None:
        return None

    oom_input = math.floor(math.log10(abs(input_value)))
    oom_reference = math.floor(math.log10(abs(reference_value)))
    
    scaling_factor = 10**(oom_input - oom_reference)
    
    scaled_value = input_value/scaling_factor
    
    return scaled_value
    

  
#%%

def Read_Request(request_directory, filename, MUSE_KRI_mapping, Bank_KRI_mapping, MUSE_sheet, Bank_sheet, formatted_date, Bank_data_index, MUSE_data_index):
    """Reads the Bank/MUSE Request files. 
    MUse/Bank_index_reg is the index[1] is for Tech metrics"""
    
    
    MUSE_index_reg = MUSE_data_index[0]
    MUSE_index_tech = MUSE_data_index[1]
    
    Bank_index_reg = Bank_data_index[0]
    Bank_index_tech = Bank_data_index[1]
    
    Tech = 0 #1 indicates its a tech submission
    
    if filename.startswith('STRAC'):
        Tech = 1
        
    request_file = os.path.join(request_directory, filename)
    request_wb = openpyxl.load_workbook(request_file)
            
    #Maybe change this to be more flexible
    
    if Tech == 1:
        request_sheet = ['KI Report_1']
        MUSE_cols = MUSE_index_tech
        Bank_cols = Bank_index_tech
    else:    
        request_sheet = request_wb['RA Request']
        MUSE_cols = MUSE_index_reg
        Bank_cols = Bank_index_reg
  
    
    for row in request_sheet.iter_rows(min_row=1, max_row=80):
        
        
        KRI, data = row[MUSE_cols[0]].value, row[MUSE_cols[1]].value
        
    
        if Tech == 1:
            KRI_string = str(KRI)
            modify = KRI_string[2:] #Take out the KI0 part
            KRI_Tech = int(modify)
            KRI = Code_Translate(KRI_Tech) #translate into RAD KRI
        
        
        if KRI in MUSE_KRI_mapping:
            target_row = MUSE_KRI_mapping[KRI]
            
        
            #Scale the data
            reference_data = MUSE_sheet.cell(row = target_row, column = 4)
            scaled_data = scale_to_order_of_magnitude(data, reference_data)
            data = scaled_data
            
                    
            MUSE_sheet.cell(row=target_row, column=4, value=data if data else None).value
            MUSE_sheet.cell(row=target_row, column=6, value=formatted_date)
            
            for row in request_sheet.iter_rows(min_row=1, max_row=99):
                KRI, data = row[Bank_cols[0]].value, row[Bank_cols[1]].value
   
                if KRI in Bank_KRI_mapping:
                    target_row = Bank_KRI_mapping[KRI]
                    Bank_sheet.cell(row=target_row, column=4, value=data if data else None)
                    
                    Bank_sheet.cell(row=target_row, column=6, value=formatted_date)
                    
                           
            request_wb.close()
    


def Update_Values(request_directory, RAD_MUSE, RAD_Bank, MUSE_data_index, Bank_data_index):
    """Calls Read Request
    Iterates through the Rows of the MUSE and Bank Dated Files and Records the KRI and the
    corresponding rows. Takes this data and udpates it in the file through Read Request"""
    
    last_business_date_str = last_business_day()
    last_business_date = datetime.datetime.strptime(last_business_date_str, '%d%b%y')
    formatted_date = last_business_date.strftime('%d/%m/%Y')
    
    MUSE_wb = openpyxl.load_workbook(RAD_MUSE)
    Bank_wb = openpyxl.load_workbook(RAD_Bank)
    
    #Dated Archive Sheet
    MUSE_sheet = MUSE_wb['ManualInputs']
    Bank_sheet = Bank_wb['ManualInputs']
    
    MUSE_KRI_mapping = {}
    Bank_KRI_mapping = {}
    
    #Iterate through the rows
    for row in MUSE_sheet.iter_rows(min_row=1, max_row=150, min_col=2, max_col=2):
        cell_value = row[0].value
   
        if cell_value:
            MUSE_KRI_mapping[cell_value] = row[0].row
    
    for row in Bank_sheet.iter_rows(min_row=1, max_row=120, min_col=2, max_col=2):
        cell_value = row[0].value
        if cell_value:
            Bank_KRI_mapping[cell_value] = row[0].row
    
 
    #go through the request directory. 
    for filename in os.listdir(request_directory):
        if filename.endswith(".xlsx"):
        
            Read_Request(request_directory, filename, MUSE_KRI_mapping, Bank_KRI_mapping, MUSE_sheet, Bank_sheet, formatted_date, MUSE_data_index, MUSE_data_index)    
        
    MUSE_wb.save(RAD_MUSE)   
    Bank_wb.save(RAD_Bank)   
    
    
    

def Run_Automation(MUSE_data_index, Bank_data_index, output_text):
    
    #base directory is location of the executable. 
    base_directory =  os.path.dirname(sys.executable)
    output_text.insert(tk.END, "Running automation...\n")
    output_text.update_idletasks()
    
    # Directory for all the risk appetite requests
    request_directory = Locate_Requests(base_directory)
    output_text.insert(tk.END, "Risk Appetite Request Located\n")
    output_text.update_idletasks()
    
    # Create the file in the dated archive and output the MUSE and Bank files needed
    MUSE_file, Bank_file = Create_Dated_RAD(base_directory)
    output_text.insert(tk.END, "Created Dated RAD files\n")
    output_text.update_idletasks()
    
    # Run the automation
    Update_Values(request_directory, MUSE_file, Bank_file, MUSE_data_index, Bank_data_index)
    output_text.insert(tk.END, "Values Updated in RADs\n")
    output_text.update_idletasks()
    
    # Process all the request files
    process_all_files(request_directory)
    output_text.insert(tk.END, "Request Files Processed\n")
    output_text.update_idletasks()
    
    # Shows procedure is done on the GUI
    output_text.insert(tk.END, "Completed\n")
    output_text.update_idletasks()
    

def main():

    
    # Create the main window
    root = tk.Tk()
    root.title("RAD Automation")
    root.configure(bg="#222222")
    
    # Create a title label
    title_label = tk.Label(root, text="MUSE/Bank RAD Automation", bg="#222222", fg="#ffffff", font=("Arial", 20, "bold"))
    title_label.pack(pady=5)
    
    # Create a label for welcome message
    welcome_label = tk.Label(root, text="Welcome to RAD Automation!", bg="#222222", fg="#cccccc", font=("Arial", 14))
    welcome_label.pack(pady=5)
    
    # Create a text widget for output message
    output_text = tk.Text(root, bg="#333333", fg="#cccccc", font=("Arial", 12), height=10, width=50)
    output_text.pack(pady=10)
    
    # Create a frame for the progress bar
    progress_frame = tk.Frame(root, bg="#222222")
    progress_frame.pack(pady=10)
    
    # Create a progress bar
    progress_bar = ttk.Progressbar(progress_frame, orient="horizontal", length=200, mode="determinate")
    progress_bar.pack(pady=5)
    
    # Function to run the code and update output message
    def update_output():
        Run_Automation(MUSE_data_index, Bank_data_index, output_text)
    
    # Define button colors
    button_bg_normal = "#4CAF50"
    button_bg_active = "#8BC34A"
    
    # Function to change button color when clicked
    def on_button_click(event):
        run_button.config(bg=button_bg_active)
    
    def on_button_release(event):
        run_button.config(bg=button_bg_normal)
        update_output()
    
    # Create a button to run the code
    run_button = tk.Button(root, text="Update RADs", bg=button_bg_normal, fg="#ffffff", font=("Arial", 16), relief=tk.FLAT)
    run_button.bind("<Button-1>", on_button_click)
    run_button.bind("<ButtonRelease-1>", on_button_release)
    run_button.config(activebackground=button_bg_active, highlightbackground=button_bg_normal, highlightcolor=button_bg_active, borderwidth=0, padx=10, pady=5)
    run_button.pack(pady=10)
    
    # Function to open the settings window
    def open_settings_window():
        settings_window = tk.Toplevel(root)
        settings_window.title("Settings")
        settings_window.configure(bg="#222222")
        
        # Create section headings in the settings window
        section1_label = tk.Label(settings_window, text="MUSE Metric Columns", bg="#222222", fg="#ffffff", font=("Arial", 14, "bold"))
        section1_label.grid(row=0, columnspan=2, padx=20, pady=10)
        
        var1_label = tk.Label(settings_window, text="MUSE KRIs:", bg="#222222", fg="#ffffff", font=("Arial", 12))
        var1_label.grid(row=1, column=0, padx=20, pady=5, sticky="e")
        var1_entry = tk.Entry(settings_window, font=("Arial", 12))
        var1_entry.grid(row=1, column=1, padx=20, pady=5)
        var1_entry.insert(0, 'A')
        
        var2_label = tk.Label(settings_window, text="MUSE Data:", bg="#222222", fg="#ffffff", font=("Arial", 12))
        var2_label.grid(row=2, column=0, padx=20, pady=5, sticky="e")
        var2_entry = tk.Entry(settings_window, font=("Arial", 12))
        var2_entry.grid(row=2, column=1, padx=20, pady=5)
        var2_entry.insert(0, 'D')
        
        section2_label = tk.Label(settings_window, text="Bank Metric Columns", bg="#222222", fg="#ffffff", font=("Arial", 14, "bold"))
        section2_label.grid(row=3, columnspan=2, padx=20, pady=10)
        
        var3_label = tk.Label(settings_window, text="Bank KRIs:", bg="#222222", fg="#ffffff", font=("Arial", 12))
        var3_label.grid(row=4, column=0, padx=20, pady=5, sticky="e")
        var3_entry = tk.Entry(settings_window, font=("Arial", 12))
        var3_entry.grid(row=4, column=1, padx=20, pady=5)
        var3_entry.insert(0, 'G')
        
        var4_label = tk.Label(settings_window, text="Bank Data:", bg="#222222", fg="#ffffff", font=("Arial", 12))
        var4_label.grid(row=5, column=0, padx=20, pady=5, sticky="e")
        var4_entry = tk.Entry(settings_window, font=("Arial", 12))
        var4_entry.grid(row=5, column=1, padx=20, pady=5)
        var4_entry.insert(0, 'K')
        
        section3_label = tk.Label(settings_window, text="Tech Metric Columns", bg="#222222", fg="#ffffff", font=("Arial", 14, "bold"))
        section3_label.grid(row=6, columnspan=2, padx=20, pady=10)
        
        var5_label = tk.Label(settings_window, text="Tech KRIs:", bg="#222222", fg="#ffffff", font=("Arial", 12))
        var5_label.grid(row=7, column=0, padx=20, pady=5, sticky="e")
        var5_entry = tk.Entry(settings_window, font=("Arial", 12))
        var5_entry.grid(row=7, column=1, padx=20, pady=5)
        var5_entry.insert(0, 'B')
        
        var6_label = tk.Label(settings_window, text="Tech Data:", bg="#222222", fg="#ffffff", font=("Arial", 12))
        var6_label.grid(row=8, column=0, padx=20, pady=5, sticky="e")
        var6_entry = tk.Entry(settings_window, font=("Arial", 12))
        var6_entry.grid(row=8, column=1, padx=20, pady=5)
        var6_entry.insert(0, 'V')
        
        # Function to handle input submission
        def submit_inputs():
            var1 = var1_entry.get()
            var2 = var2_entry.get()
            var3 = var3_entry.get()
            var4 = var4_entry.get()
            var5 = var5_entry.get()
            var6 = var6_entry.get()
            
            # Convert letters to numbers (A=0, B=1, etc.)
            global MUSE_data_index, Bank_data_index
            MUSE_data_index = [[ord(var1.upper()) - 65, ord(var2.upper()) - 65], [ord(var5.upper()) - 65, ord(var6.upper()) - 65]]
            Bank_data_index = [[ord(var3.upper()) - 65, ord(var4.upper()) - 65], [ord(var5.upper()) - 65, ord(var6.upper()) - 65]]
            
            output_text.insert(tk.END, f"Submitted Columns\n")
            output_text.update_idletasks()
            settings_window.destroy()  # Close the settings window
        
        # Create a button to submit the inputs
        submit_button = tk.Button(settings_window, text="Submit Variables", bg=button_bg_normal, fg="#ffffff", font=("Arial", 16), relief=tk.FLAT, command=submit_inputs)
        submit_button.grid(row=9, columnspan=2, pady=10)
    
    # Create a button to open the settings window
    settings_button = tk.Button(root, text="Settings", bg="#2196F3", fg="#ffffff", font=("Arial", 16), relief=tk.FLAT, command=open_settings_window)
    settings_button.pack(pady=10)
    
    # Start the Tkinter event loop
    root.mainloop()

if __name__ == "__main__":
    main()
